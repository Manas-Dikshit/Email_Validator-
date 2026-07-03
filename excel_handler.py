"""
excel_handler.py
----------------
Reads email data from the input Excel file and writes validation
results back to the output files.

Expected input columns (case-insensitive):
  * Mail ID     - the email address
  * Mail Status - optional existing status (will be overwritten)

Output adds:
  * Mail Status       - validation status code
  * Validation Reason - human-readable explanation
  * SMTP Response     - raw SMTP code + message
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    FAILED_EXCEL,
    OUTPUT_DIR,
    OUTPUT_EXCEL,
    STATUS_VALID,
)
from logger import get_logger

log = get_logger(__name__)

PathLike = Union[str, os.PathLike]

# Canonical column order/names used for every output workbook.
_RESULT_KEYS = ["email", "status", "reason", "smtp_response"]
_OUTPUT_HEADERS = ["Mail ID", "Mail Status", "Validation Reason", "SMTP Response"]

# ── Colour palette for the status column ─────────────────────────────────────
_STATUS_COLOURS = {
    "VALID":             "C6EFCE",  # green
    "INVALID_FORMAT":    "FFEB9C",  # yellow
    "INVALID_DOMAIN":    "FFC7CE",  # red-light
    "INVALID_MAILBOX":   "FFC7CE",  # red-light
    "ACCESS_DENIED":     "FFCC99",  # orange
    "CATCH_ALL":         "BDD7EE",  # blue-light
    "TEMPORARY_FAILURE": "E2EFDA",  # grey-green
    "UNKNOWN":           "D9D9D9",  # grey
}
_DEFAULT_STATUS_COLOUR = "FFFFFF"


def read_emails(filepath: PathLike, dedupe: bool = True) -> List[str]:
    """
    Read email addresses from an Excel file.

    Looks for a column whose name contains 'mail id' (case-insensitive).
    Falls back to the first column if no match is found.

    Parameters
    ----------
    filepath : str | os.PathLike
        Path to the .xlsx input file.
    dedupe : bool
        If True (default), remove duplicate addresses (case-insensitive)
        while preserving first-seen order, so the pipeline doesn't waste
        SMTP round-trips re-validating the same address twice.

    Returns
    -------
    list[str]
        All non-empty, whitespace-trimmed email strings found in the column.

    Raises
    ------
    FileNotFoundError
        If filepath does not exist.
    ValueError
        If the file has no columns / is empty.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Input file not found: {filepath}")

    log.info("Reading input file: %s", filepath)
    try:
        df = pd.read_excel(filepath, dtype=str)
    except Exception as exc:  # noqa: BLE001 - surface a clear, actionable error
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    if df.empty or len(df.columns) == 0:
        raise ValueError(f"Input file '{filepath}' has no columns / is empty")

    # Find the email column
    col = _find_column(df, "mail id") or df.columns[0]
    log.info("Using column '%s' for email addresses", col)

    emails = df[col].dropna().astype(str).str.strip().tolist()
    emails = [e for e in emails if e and e.lower() != "nan"]  # drop blanks/stringified NaN

    total_found = len(emails)

    if dedupe:
        seen = set()
        unique_emails = []
        for e in emails:
            key = e.lower()
            if key not in seen:
                seen.add(key)
                unique_emails.append(e)
        emails = unique_emails
        if len(emails) != total_found:
            log.info(
                "Removed %d duplicate address(es); %d unique remain",
                total_found - len(emails), len(emails),
            )

    log.info("Found %d email address(es)", len(emails))
    return emails


def write_results(results: List[dict]) -> Optional[Path]:
    """
    Write validation results to the output Excel file with colour-coding.

    Parameters
    ----------
    results : list[dict]
        Each dict should have keys: email, status, reason, smtp_response.
        Missing keys are treated as blank; extra keys are ignored.

    Returns
    -------
    Optional[Path]
        Path to the written file, or None if there was nothing to write.
    """
    if not results:
        log.warning("write_results called with an empty results list - nothing to save")
        return None

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    df = _results_to_dataframe(results)
    wb = _build_workbook(df, sheet_name="Validation Results")

    out_path = Path(OUTPUT_EXCEL)
    wb.save(out_path)
    log.info("Results saved -> %s (%d rows)", out_path, len(df))
    return out_path


def write_failed_emails(results: List[dict]) -> Optional[Path]:
    """
    Write only the failed / invalid email results to a separate file.

    Parameters
    ----------
    results : list[dict]
        Full results list; this function filters out VALID emails.

    Returns
    -------
    Optional[Path]
        Path to the written file, or None if there were no failures
        (or no results at all).
    """
    if not results:
        log.warning("write_failed_emails called with an empty results list")
        return None

    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    failed = [r for r in results if r.get("status") != STATUS_VALID]
    if not failed:
        log.info("No failed emails - skipping failed_emails.xlsx")
        return None

    df = _results_to_dataframe(failed)
    wb = _build_workbook(df, sheet_name="Failed Emails")

    out_path = Path(FAILED_EXCEL)
    wb.save(out_path)
    log.info("Failed emails saved -> %s (%d rows)", out_path, len(df))
    return out_path


# ── Private helpers ───────────────────────────────────────────────────────────

def _results_to_dataframe(results: List[dict]) -> pd.DataFrame:
    """
    Normalise a list of result dicts into a DataFrame with the canonical
    output columns, regardless of missing/extra keys or None values.
    """
    df = pd.DataFrame(results)
    # reindex guarantees every expected column exists (filled with "" if
    # absent) and drops anything not in _RESULT_KEYS (e.g. mx_hosts).
    df = df.reindex(columns=_RESULT_KEYS, fill_value="")
    df = df.fillna("")
    df.columns = _OUTPUT_HEADERS
    return df


def _find_column(df: pd.DataFrame, keyword: str) -> Optional[str]:
    """
    Return the best-matching column name for `keyword`.

    Prefers an exact (case-insensitive) match, falling back to the first
    column whose name merely contains the keyword as a substring.
    """
    keyword_lower = keyword.lower()
    substring_match: Optional[str] = None

    for col in df.columns:
        col_lower = str(col).strip().lower()
        if col_lower == keyword_lower:
            return col
        if substring_match is None and keyword_lower in col_lower:
            substring_match = col

    return substring_match


def _build_workbook(df: pd.DataFrame, sheet_name: str) -> Workbook:
    """
    Convert a DataFrame into a formatted openpyxl Workbook.

    Parameters
    ----------
    df : pd.DataFrame
        Data to write (already has correct column names).
    sheet_name : str
        Name for the worksheet (truncated to Excel's 31-char limit).

    Returns
    -------
    Workbook
        Ready-to-save workbook with header formatting, row colours,
        frozen header row, and an autofilter.
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = sheet_name[:31]  # Excel sheet names are capped at 31 chars

    # ── Headers ───────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill  = PatternFill("solid", start_color="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Data rows ─────────────────────────────────────────────────────────────
    columns = list(df.columns)
    status_col_idx = columns.index("Mail Status") + 1 if "Mail Status" in columns else None

    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            # Guard against stray floats/NaN slipping through (e.g. pandas
            # upcasting an all-blank column) so cells never render "nan".
            if value is None or (isinstance(value, float) and pd.isna(value)):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align

        if status_col_idx is not None:
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            colour = _STATUS_COLOURS.get(str(status_cell.value or ""), _DEFAULT_STATUS_COLOUR)
            status_cell.fill = PatternFill("solid", start_color=colour)

    # ── Column widths ─────────────────────────────────────────────────────────
    default_widths = [35, 22, 45, 20]
    for i in range(1, len(columns) + 1):
        width = default_widths[i - 1] if i <= len(default_widths) else 25
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20

    # ── Usability: freeze header row + enable filtering/sorting ──────────────
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    return wb